
import unittest
import os
import sys
from openpyxl import Workbook
from unittest.mock import MagicMock, patch

# Add repo root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    from skills.xlsx.recalc import count_formulas
except ImportError:
    count_formulas = None

class TestRecalcPerf(unittest.TestCase):
    def setUp(self):
        self.filename = "test_perf_formulas.xlsx"
        self.create_test_file(self.filename)

    def tearDown(self):
        if os.path.exists(self.filename):
            os.remove(self.filename)

    def create_test_file(self, filename, rows=100, cols=10):
        wb = Workbook()
        ws = wb.active

        # Add formulas
        # A1 to A{rows} will have formulas
        for r in range(1, rows + 1):
            ws[f'A{r}'] = f'=B{r}+C{r}'

        # Add values (no formula)
        for r in range(1, rows + 1):
            ws[f'B{r}'] = 1
            ws[f'C{r}'] = 2

        wb.save(filename)
        # Expected formulas: rows
        self.expected_count = rows

    def test_count_formulas(self):
        if count_formulas is None:
            # If function doesn't exist, we can't test it.
            # This allows creating the test file before the implementation.
            print("count_formulas not found, skipping test")
            return

        count = count_formulas(self.filename)
        self.assertEqual(count, self.expected_count, f"Expected {self.expected_count} formulas, got {count}")

    def test_small_chunks(self):
        """Test that formula counting works even when file is read in small chunks (splitting tags)"""
        if count_formulas is None:
             self.skipTest("count_formulas not implemented yet")

        # XML content with formulas split in various ways
        # <f> in simple form
        # <f t="shared"> with attributes
        # <f> split across chunk boundary
        xml_content = b'<worksheet><sheetData><row><c><f>SUM(A1)</f></c><c><f t="shared" ref="A1">SUM(A1)</f></c><c><f>A1+1</f></c></row></sheetData></worksheet>'
        # Total 3 formulas.

        # Mocking zipfile
        with patch('zipfile.ZipFile') as MockZip:
            mock_zip = MockZip.return_value
            mock_zip.__enter__.return_value = mock_zip
            mock_zip.namelist.return_value = ['xl/worksheets/sheet1.xml']

            mock_file = MagicMock()
            mock_file.__enter__.return_value = mock_file
            mock_zip.open.return_value = mock_file

            # Use small chunk size to force splitting (e.g., 5 bytes)
            chunks = []
            chunk_size = 5
            for i in range(0, len(xml_content), chunk_size):
                chunks.append(xml_content[i:i+chunk_size])

            iter_chunks = iter(chunks)

            def side_effect(size=-1):
                try:
                    return next(iter_chunks)
                except StopIteration:
                    return b''

            mock_file.read.side_effect = side_effect

            # Filename doesn't matter as we mock ZipFile
            count = count_formulas("dummy.xlsx")
            self.assertEqual(count, 3)

if __name__ == '__main__':
    unittest.main()
